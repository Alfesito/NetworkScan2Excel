from scapy.all import ARP, Ether, srp
from mac_vendor_lookup import MacLookup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
import os, nmap

def escanear_red_arp(red, intentos=5, timeout=5):
    """
    Escanea la red utilizando ARP, combina resultados únicos y usa map para procesar los datos.
    """
    dispositivos = {}

    for intento in range(intentos):
        paquete = Ether(dst="ff:ff:ff:ff:ff:ff") / ARP(pdst=red)
        respuestas = srp(paquete, timeout=timeout, verbose=False)[0]

        def procesar_respuesta(par):
            _, recibido = par
            mac = recibido.hwsrc
            ip = recibido.psrc
            try:
                proveedor = MacLookup().lookup(mac)
            except Exception:
                proveedor = "Proveedor desconocido"
            return mac, {"IP": ip, "Proveedor": proveedor}

        dispositivos_actualizados = dict(map(procesar_respuesta, respuestas))
        dispositivos.update(dispositivos_actualizados)

    return [{"MAC": mac, "IP": datos["IP"], "Proveedor": datos["Proveedor"]}
            for mac, datos in dispositivos.items()]

def detectar_sistema_operativo(ips):
    """
    Detecta el sistema operativo de una lista de IPs utilizando Nmap.
    """
    nm = nmap.PortScanner()
    os_detected = {}

    for ip in ips:
        try:
            scan = nm.scan(hosts=ip, arguments="-O -n -Pn --min-rate 5000", sudo=True)
            os_info = scan["scan"].get(ip, {}).get("osmatch", [])
            os_detected[ip] = os_info[0]["name"] if os_info else "Desconocido"
        except Exception as e:
            os_detected[ip] = "Error/No detectado"

    return os_detected

def guardar_en_excel(dispositivos, archivo_excel, red_objetivo):
    """
    Guarda la información de los dispositivos en un archivo Excel.
    Si ya existe, compara los datos y resalta cambios en amarillo.
    Si algún dispositivo ya no está presente, se marca en rojo.
    """
    nombre_hoja = red_objetivo.replace("/", "-")
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    sin_color = PatternFill(fill_type=None)

    archivo_existe = True

    if os.path.exists(archivo_excel):
        wb = load_workbook(archivo_excel)
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
        else:
            ws = wb.create_sheet(title=nombre_hoja)
            archivo_existe = False
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = nombre_hoja
        archivo_existe = False

    datos_existentes = set()
    for fila in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3, values_only=True):
        ip, mac = fila
        if ip and mac:
            datos_existentes.add((ip, mac))

    nuevos_dispositivos = []
    for dispositivo in dispositivos:
        nuevos_dispositivos.append({
            'IP': dispositivo['IP'],
            'MAC': dispositivo['MAC'],
            'Proveedor': dispositivo['Proveedor'],
            'OS': dispositivo.get('OS', "Desconocido")
        })

    dispositivos_a_agregar = [d for d in nuevos_dispositivos if (d['IP'], d['MAC']) not in datos_existentes]
    dispositivos_a_eliminar = [(ip, mac) for (ip, mac) in datos_existentes if not any(d['IP'] == ip and d['MAC'] == mac for d in nuevos_dispositivos)]

    if ws.max_row == 1:
        ws['B1'] = "IP"
        ws['C1'] = "MAC"
        ws['E1'] = "Proveedor"
        ws['F1'] = "OS"

    if archivo_existe:
        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3), start=2):
            ip, mac = fila[0].value, fila[1].value
            if (ip, mac) in dispositivos_a_eliminar:
                ws[f'B{fila_num}'].fill = rojo
                ws[f'C{fila_num}'].fill = rojo

    fila = ws.max_row + 1
    for dispositivo in dispositivos_a_agregar:
        ip = dispositivo['IP']
        mac = dispositivo['MAC']
        proveedor = dispositivo['Proveedor']
        os_info = dispositivo['OS']
        
        ws[f'B{fila}'] = ip
        ws[f'C{fila}'] = mac
        ws[f'E{fila}'] = proveedor
        ws[f'F{fila}'] = os_info

        if archivo_existe:
            ws[f'B{fila}'].fill = amarillo
            ws[f'C{fila}'].fill = amarillo
        fila += 1

    wb.save(archivo_excel)
    print(f"Datos guardados en {archivo_excel}")

def main():
    red_objetivo = input("Introduce el rango de red (por ejemplo, 192.168.1.0/24): ").strip()
    archivo_excel = input("Introduce el nombre del archivo Excel de salida (por ejemplo, dispositivos_red.xlsx): ").strip()

    print(f"Escaneando la red: {red_objetivo}...")
    dispositivos = escanear_red_arp(red_objetivo)
    ips = [d["IP"] for d in dispositivos]

    print("Detectando sistemas operativos...")
    os_detectados = detectar_sistema_operativo(ips)

    for dispositivo in dispositivos:
        dispositivo["OS"] = os_detectados.get(dispositivo["IP"], "Desconocido")

    if dispositivos:
        print(f"Se encontraron {len(dispositivos)} dispositivos únicos.")
        guardar_en_excel(dispositivos, archivo_excel, red_objetivo)
    else:
        print("No se encontraron dispositivos en la red.")

if __name__ == "__main__":
    main()
